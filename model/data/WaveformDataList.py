from numpy import ndarray, array
from typing import Iterable
from openpyxl import Workbook, load_workbook
import pickle


class WaveformData:
    def __init__(self):
        self.f = None
        self.f_unit = "Hz"
        self.t = None
        self.t_unit = "s"
        self.v = None
        self.v_unit = "mV"
        self.v_count = 0
        self.r_abs = None
        self.r_unit = "ohm"
        self.r_int = None

    def set_data(self, f: float, t: ndarray, v: ndarray, r: float, ri: float):
        self.f = f
        self.t = t
        self.v = v
        self.v_count = len(v)
        self.r_abs = r
        self.r_int = ri

    def set_constants(self, f: float, r: float):
        self.f = f
        self.r_abs = r

    def get_time_count(self):
        return len(self.t)

    def get_v_count(self):
        return self.v_count

    def get_data(self):
        return [self.t, self.v]

    def get_frequency(self) -> float:
        return self.f

    def get_time(self):
        return self.t

    def get_v(self, index):
        return array(self.v[index])

    def get_measuring_resistance(self) -> float:
        return self.r_abs

    def get_measuring_r_int(self) -> float:
        return self.r_int

    def save_data(self, filename):
        if not filename.endswith('xlsx'):
            filename += ".xlsx"

        wb = Workbook()
        ws = wb.create_sheet("Waveform", 0)

        ws.cell(1, 1, "Requested Frequency")
        ws.cell(1, 2, str(self.f))
        ws.cell(1, 2).number_format = '0.0000'
        ws.cell(1, 3, self.f_unit)

        ws.cell(2, 1, "Current Resistor")
        ws.cell(2, 2, str(self.r_abs))
        ws.cell(2, 2).number_format = '0.00'
        ws.cell(2, 3, "ohm")

        ws.cell(3, 1, "X9C103S Int")
        ws.cell(3, 2, str(self.r_int))
        ws.cell(3, 2).number_format = '0'
        ws.cell(3, 3, "unitless")

        ws.cell(4, 1, 't')
        ws.cell(5, 1, self.t_unit)
        for i in range(self.v_count):
            ws.cell(4, 2 + i, "V{0}".format(i))
            ws.cell(5, 2 + i, "volts")

        n = len(self.t)
        for i in range(n):
            ws.cell(6 + i, 1, self.t[i])
            for j in range(self.v_count):
                ws.cell(6 + i, 2 + j, self.v[j][i])

        wb.save(filename)

    def load_data(self, filename):
        if not filename.endswith('xlsx'):
            filename += ".xlsx"
        wb = load_workbook(filename)
        ws = wb["Waveform"]
        i = 0
        self.t = []
        self.v = []
        for row in ws.values:
            if i == 0:
                self.f = float(row[1])
                self.f_unit = str(row[2])
            if i == 1:
                self.r_abs = float(row[1])
                self.r_unit = str(row[2])
            if i == 2:
                if not row[1] is None:
                    self.r_int = float(row[1])

            if i == 4:
                self.t_unit = str(row[0])
                self.v_unit = str(row[1])
                for j in range(len(row)-1):
                    self.v.append([])

            if i > 4:
                self.t.append(float(row[0]))
                for j in range(len(row)-1):
                    self.v[j].append(float(row[j+1]))
            i += 1

        self.v_count = len(self.v)

class WaveformDataList:
    def __init__(self):
        self.wfs = []

    def append(self, wf: WaveformData):
        self.wfs.append(wf)

    def at_index(self, index: int) -> WaveformData:
        if index < len(self.wfs):
            return self.wfs[index]

    def remove_at_index(self, index: int):
        self.wfs.remove(self.wfs[index])

    def get_count(self):
        return len(self.wfs)

    def save_waveforms(self, filename):
        my_file = open(filename, 'wb')
        pickle.dump(self.wfs, my_file, pickle.HIGHEST_PROTOCOL)

    def load_waveforms(self, filename: str):
        if not filename.endswith('pkl'):
            filename += '.pkl'
        my_file = open(filename, 'rb')
        self.wfs = pickle.load(my_file)

# wd = WaveformData()
# #wd.set_data(10, array([1, 2, 3, 4, 5, 6, 7, 8, 9, 10]), array([[1,2,3,4,5,6,7,8,9,0], [10,11,12,13,14,15,16,17,18,19]]), 330, 0)
# #wd.save_data("testfile")
# wd.load_data("testfile")
# print(wd.t)
# print(wd.v)