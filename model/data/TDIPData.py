from numpy import ndarray, zeros, array, exp, real, imag
from openpyxl import Workbook, load_workbook
from typing import Iterable

# A class to hold the time domain ip data.
class TDIPData:
    def __init__(self, mr):
        self.t = []
        self.Va = []
        self.Vm = []
        self.Vn = []
        self.Vb = []
        self.mr = mr

    def get_time(self):
        return self.t

    def get_current(self):
        return array(self.Va) / array(self.mr)

    def get_potential_difference(self):
        return array(self.Vm) - array(self.Vn)

    def get_potential_m(self):
        return array(self.Vm)

    def get_potential_n(self):
        return array(self.Vn)

    def load_data_from_file(self, filename):
        if not filename.endswith('xlsx'):
            filename += "_Z.xlsx"
        wb = load_workbook(filename)
        ws = wb["ImpedanceData"]
        i = 0
        for row in ws.values:
            if i > 1:
                self.t.append(float(row[0]))
                self.Va.append(float(row[1]))
                self.Vm.append(float(row[2]))
                self.Vn.append(float(row[3]))
                self.Vb.append(float(row[4]))
            i += 1

    def save_data_to_file(self, filename: str):
        if not filename.endswith('xlsx'):
            filename += '_Z.xlsx'
        wb = Workbook()
        ws = wb.create_sheet("ImpedanceData", 0)

        ws.cell(1, 1, "t")
        ws.cell(2, 1, "s")

        ws.cell(1, 2, "Va")
        ws.cell(2, 2, "V")

        ws.cell(1, 3, "Vm")
        ws.cell(2, 3, "V")

        ws.cell(1, 4, "Vn")
        ws.cell(2, 4, "V")

        ws.cell(1, 5, "Vb")
        ws.cell(2, 5, "V")

        n = len(self.t)
        if n > 1:
            for i in range(n):
                ws.cell(i + 3, 1, self.t[i])
                ws.cell(i + 3, 1).number_format = '0.00'
                ws.cell(i + 3, 2, self.Va[i])
                ws.cell(i + 3, 2).number_format = '0.00'
                ws.cell(i + 3, 3, self.Vm[i])
                ws.cell(i + 3, 3).number_format = '0.00'
                ws.cell(i + 3, 4, self.Vn[i])
                ws.cell(i + 3, 4).number_format = '0.00'
                ws.cell(i + 3, 5, self.Vb[i])
                ws.cell(i + 3, 5).number_format = '0.00'

        wb.save(filename)

    def set_data(self, t: ndarray, va: ndarray, vm: ndarray, vn: ndarray, vb: ndarray):
        self.t = t
        self.Va = va
        self.Vm = vm
        self.Vn = vn
        self.Vb = vb

