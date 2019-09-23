from numpy import ndarray, zeros, array, exp, real, imag
from openpyxl import Workbook, load_workbook
from typing import Iterable, List

class FitData:
    def __init__(self, n_param: int):
        self.n_param = n_param
        self.i = []
        self.a = []
        self.f = []
        self.p = []
        self.d = []
        self.e = []

    def set_data(self, i: List[int], a: List[float], f: List[float], p: List[float], d: List[float], e: List[float]):
        self.i = i
        self.a = a
        self.f = f
        self.p = p
        self.d = d

    def append_data_from_array(self, i: int, parameters: Iterable, e: float):
        self.i.append(i)
        a, f, p, d = parameters

        self.a.append(a)
        self.f.append(f)
        self.p.append(p)
        self.d.append(d)
        self.e.append(e)

    def get_length(self):
        return int(len(self.i) / self.n_param)

    def get_data(self, index, i: int):
        index = self.n_param*index + i
        return [self.i[index], self.a[index], self.f[index], self.p[index], self.d[index], self.e[index]]

    def save_data(self, filename: str):
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        wb = Workbook()
        ws = wb.create_sheet("VoltageData")

        ws.cell(1, 1, "WF Index")
        ws.cell(1, 2, "A")
        ws.cell(1, 3, "F")
        ws.cell(1, 4, "P")
        ws.cell(1, 5, "D")
        ws.cell(1, 6, "Error")

        n = len(self.i)
        for i in range(n):
            ws.cell(i + 2, 1, self.i[i])
            ws.cell(i + 2, 2, self.a[i])
            ws.cell(i + 2, 3, self.f[i])
            ws.cell(i + 2, 4, self.p[i])
            ws.cell(i + 2, 5, self.d[i])
            ws.cell(i + 2, 6, self.e[i])

        wb.save(filename)

    def load_data(self, filename: str):
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        wb = load_workbook(filename)
        ws = wb["VoltageData"]

        i = 1
        for row in ws.values:
            if i > 1:
                self.i.append(int(row[0]))
                self.a.append(float(row[1]))
                self.f.append(float(row[2]))
                self.p.append(float(row[3]))
                self.d.append(float(row[4]))
                self.e.append(float(row[5]))
            i += 1
